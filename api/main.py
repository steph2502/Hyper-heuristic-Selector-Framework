"""FastAPI backend for timetable optimization and exports."""

from __future__ import annotations

import csv
import os
import sys
import time
import uuid
from pathlib import Path
from typing import Any

from fastapi import Body, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse

API_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = API_DIR.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from algorithms.aco import run_aco
from algorithms.controller import run_hyper_heuristic
from algorithms.fitness import evaluate_timetable
from algorithms.initializer import count_scheduled_lectures, generate_initial_solution
from algorithms.pso import run_pso
from models.course import Course
from models.room import Room
from models.timetable import TimetableState
from parsers.itc_parser import (
    Curriculum,
    ITCInstance,
    UnavailabilityConstraint,
    parse_itc_file,
)
from utils.timetable_display import getDayName, getTimeSlot
from utils.timetable_output import export_timetable_csv

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment]

IS_VERCEL = os.environ.get("VERCEL") == "1"
# Vercel functions can only write to /tmp at runtime.
RUNTIME_DIR = Path("/tmp") if IS_VERCEL else API_DIR
UPLOAD_DIR = RUNTIME_DIR / "uploads"
OUTPUT_DIR = RUNTIME_DIR / "outputs"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

ALGORITHMS = {"greedy", "aco", "pso", "controller"}
REQUIRED_COURSE_COLUMNS = (
    "course_code",
    "course_title",
    "lecturer_id",
    "lectures_per_week",
    "min_working_days",
    "number_of_students",
    "department",
    "level",
)
HEADER_ALIASES: dict[str, set[str]] = {
    "course_code": {"course_code", "course_code_id", "coursecode", "course", "course_id", "code"},
    "course_title": {"course_title", "coursetitle", "title", "course_name", "name"},
    "lecturer_id": {"lecturer_id", "lecturer", "teacher_id", "teacher", "instructor_id", "instructor"},
    "lectures_per_week": {
        "lectures_per_week",
        "lecture_per_week",
        "lectures_week",
        "lectures_each_week",
        "weekly_lectures",
        "sessions_per_week",
    },
    "min_working_days": {
        "min_working_days",
        "minimum_working_days",
        "min_days",
        "minimum_days",
    },
    "number_of_students": {
        "number_of_students",
        "students",
        "student_count",
        "enrollment",
        "enrolment",
    },
    "department": {"department", "dept"},
    "level": {"level", "academic_level", "class_level", "year"},
    "lecturer_name": {"lecturer_name", "lecturer_full_name", "teacher_name", "instructor_name"},
    "room_id": {"room_id", "room"},
    "room_name": {"room_name", "room_title"},
    "capacity": {"capacity", "room_capacity"},
    "curriculum_id": {"curriculum_id", "curriculum", "group_id"},
    "unavailability_day": {"unavailability_day", "unavailable_day", "day_unavailable"},
    "unavailability_period": {"unavailability_period", "unavailable_period", "period_unavailable"},
}
HEADER_ALIAS_LOOKUP: dict[str, str] = {
    alias: canonical for canonical, aliases in HEADER_ALIASES.items() for alias in aliases
}

app = FastAPI(title="University Timetabling Optimization API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/")
def health() -> dict[str, str]:
    """Basic status endpoint."""
    return {"message": "University Timetabling Optimization API is running."}


@app.post("/optimize")
async def optimize(
    algorithm: str = Form(...),
    file: UploadFile | None = File(default=None),
    raw_text: str | None = Form(default=None),
) -> dict[str, Any]:
    """Run one optimizer on uploaded/raw dataset and return results."""
    normalized_algo = algorithm.strip().lower()
    if normalized_algo not in ALGORITHMS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported algorithm '{algorithm}'. Use one of: {sorted(ALGORITHMS)}.",
        )

    if file is None and not (raw_text and raw_text.strip()):
        raise HTTPException(
            status_code=400,
            detail="Provide either a dataset file or non-empty raw_text.",
        )

    if file is not None and _is_legacy_excel(file.filename or ""):
        raise HTTPException(
            status_code=400,
            detail=(
                "Legacy Excel .xls files are not supported yet. "
                "Please re-save the sheet as .xlsx and upload again."
            ),
        )

    if file is not None and _looks_like_excel(file.filename or ""):
        if load_workbook is None:
            raise HTTPException(
                status_code=500,
                detail="Excel parsing support is unavailable. Install openpyxl.",
            )
        try:
            content = await file.read()
            school_payload = _school_payload_from_excel_bytes(content, file.filename or "uploaded.xlsx")
            school_payload["algorithm"] = normalized_algo
            instance, course_meta = _instance_from_school_data(school_payload)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=f"Failed to parse dataset: {exc}") from exc
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=400, detail=f"Failed to parse dataset: {exc}") from exc

        started = time.perf_counter()
        try:
            state, history, selected = _run_optimizer(normalized_algo, instance)
            evaluate_timetable(state, instance)
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=500, detail=f"Optimization failed: {exc}") from exc

        elapsed = time.perf_counter() - started
        scheduled, total = count_scheduled_lectures(state)
        output_filename = _build_output_filename(instance.name, normalized_algo)
        output_path = OUTPUT_DIR / output_filename
        _export_school_timetable_csv(state, output_path, course_meta)
        return {
            "dataset_name": instance.name,
            "algorithm": normalized_algo,
            "final_fitness": float(state.fitness or 0.0),
            "hard_violations": int(state.hard_violations or 0),
            "soft_penalty": float(state.soft_penalty or 0.0),
            "runtime_seconds": round(elapsed, 4),
            "scheduled_lectures": scheduled,
            "total_lectures": total,
            "convergence_history": [float(x) for x in history],
            "selected_heuristics": selected,
            "timetable_rows": _serialize_timetable_rows(state, course_meta),
            "download_url": f"/download/{output_filename}",
        }

    dataset_path = await _save_dataset_input(file, raw_text)
    try:
        instance = parse_itc_file(dataset_path)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Failed to parse dataset: {exc}") from exc

    started = time.perf_counter()
    try:
        state, history, selected = _run_optimizer(normalized_algo, instance)
        evaluate_timetable(state, instance)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Optimization failed: {exc}") from exc

    elapsed = time.perf_counter() - started
    scheduled, total = count_scheduled_lectures(state)

    output_filename = _build_output_filename(instance.name or dataset_path.stem, normalized_algo)
    output_path = OUTPUT_DIR / output_filename
    export_timetable_csv(state, instance, output_path)
    course_meta = _course_meta_from_instance(instance)

    return {
        "dataset_name": instance.name or dataset_path.stem,
        "algorithm": normalized_algo,
        "final_fitness": float(state.fitness or 0.0),
        "hard_violations": int(state.hard_violations or 0),
        "soft_penalty": float(state.soft_penalty or 0.0),
        "runtime_seconds": round(elapsed, 4),
        "scheduled_lectures": scheduled,
        "total_lectures": total,
        "convergence_history": [float(x) for x in history],
        "selected_heuristics": selected,
        "timetable_rows": _serialize_timetable_rows(state, course_meta),
        "download_url": f"/download/{output_filename}",
    }


@app.post("/optimize-school-data")
def optimize_school_data(payload: dict[str, Any] = Body(...)) -> dict[str, Any]:
    """Run one optimizer on structured school data entered from forms."""
    algorithm_value = str(payload.get("algorithm", "")).strip().lower()
    if algorithm_value not in ALGORITHMS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported algorithm '{payload.get('algorithm')}'. Use one of: {sorted(ALGORITHMS)}.",
        )

    try:
        instance, course_meta = _instance_from_school_data(payload)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    started = time.perf_counter()
    try:
        state, history, selected = _run_optimizer(algorithm_value, instance)
        evaluate_timetable(state, instance)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Optimization failed: {exc}") from exc

    elapsed = time.perf_counter() - started
    scheduled, total = count_scheduled_lectures(state)

    output_filename = _build_output_filename(instance.name, algorithm_value)
    output_path = OUTPUT_DIR / output_filename
    _export_school_timetable_csv(state, output_path, course_meta)

    return {
        "dataset_name": instance.name,
        "algorithm": algorithm_value,
        "final_fitness": float(state.fitness or 0.0),
        "hard_violations": int(state.hard_violations or 0),
        "soft_penalty": float(state.soft_penalty or 0.0),
        "runtime_seconds": round(elapsed, 4),
        "scheduled_lectures": scheduled,
        "total_lectures": total,
        "convergence_history": [float(x) for x in history],
        "selected_heuristics": selected,
        "timetable_rows": _serialize_timetable_rows(state, course_meta),
        "download_url": f"/download/{output_filename}",
    }


@app.get("/download/{filename}")
def download(filename: str) -> FileResponse:
    """Download a generated CSV timetable."""
    if not filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only CSV files are downloadable.")

    candidate = (OUTPUT_DIR / filename).resolve()
    output_root = OUTPUT_DIR.resolve()
    if output_root not in candidate.parents or not candidate.exists():
        raise HTTPException(status_code=404, detail="File not found.")

    return FileResponse(candidate, media_type="text/csv", filename=filename)


def _run_optimizer(
    algorithm: str,
    instance: ITCInstance,
) -> tuple[TimetableState, list[float], list[str]]:
    if algorithm == "greedy":
        state = generate_initial_solution(instance)
        evaluate_timetable(state, instance)
        return state, [float(state.fitness or 0.0)], []
    if algorithm == "aco":
        state, history = run_aco(instance, verbose=False)
        return state, history, []
    if algorithm == "pso":
        state, history = run_pso(instance, verbose=False)
        return state, history, []

    state, history, stats = run_hyper_heuristic(instance)
    return state, history, list(stats["selection_history"])


async def _save_dataset_input(file: UploadFile | None, raw_text: str | None) -> Path:
    token = uuid.uuid4().hex[:12]
    if file is not None:
        file_name = Path(file.filename or "uploaded_dataset.txt").name
        target = UPLOAD_DIR / f"{token}_{file_name}"
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="Uploaded file is empty.")
        target.write_bytes(content)
        return target

    text = (raw_text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="raw_text is empty.")
    target = UPLOAD_DIR / f"{token}_raw_dataset.txt"
    target.write_text(text, encoding="utf-8")
    return target


def _build_output_filename(dataset_name: str, algorithm: str) -> str:
    safe = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in dataset_name)
    safe = safe.strip("_") or "dataset"
    suffix = uuid.uuid4().hex[:8]
    return f"{safe}_{algorithm}_{suffix}.csv"


def _looks_like_excel(filename: str) -> bool:
    lower = filename.lower()
    return lower.endswith(".xlsx") or lower.endswith(".xlsm") or lower.endswith(".xltx")


def _is_legacy_excel(filename: str) -> bool:
    return filename.lower().endswith(".xls")


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lstrip("\ufeff").lower()
    normalized = "".join(ch if ch.isalnum() else "_" for ch in text)
    while "__" in normalized:
        normalized = normalized.replace("__", "_")
    return normalized.strip("_")


def _header_index_from_row(row: tuple[Any, ...]) -> dict[str, int]:
    header_index: dict[str, int] = {}
    for idx, value in enumerate(row):
        normalized = _normalize_header(value)
        if not normalized:
            continue
        canonical = HEADER_ALIAS_LOOKUP.get(normalized)
        if canonical and canonical not in header_index:
            header_index[canonical] = idx
    return header_index


def _safe_cell(values: list[Any], index: int | None) -> Any:
    if index is None:
        return None
    if index < 0 or index >= len(values):
        return None
    return values[index]


def _safe_int(value: Any, default: int, field_name: str, row_number: int) -> int:
    if value is None:
        return default
    text = str(value).strip()
    if not text:
        return default
    try:
        return int(float(text))
    except ValueError as exc:
        raise ValueError(
            f"Invalid numeric value '{value}' for '{field_name}' at Excel row {row_number}."
        ) from exc


def _detect_course_header(
    wb: Any,
) -> tuple[list[tuple[Any, ...]], dict[str, int], str, int]:
    best_score = -1
    best_sheet = ""
    best_row_number = 0
    best_columns: list[str] = []
    required_count = len(REQUIRED_COURSE_COLUMNS)

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        inspect_limit = min(30, len(rows))
        for row_idx in range(inspect_limit):
            row = rows[row_idx] or ()
            header_index = _header_index_from_row(row)
            score = sum(1 for col in REQUIRED_COURSE_COLUMNS if col in header_index)

            if score > best_score:
                best_score = score
                best_sheet = str(sheet.title)
                best_row_number = row_idx + 1
                best_columns = sorted(header_index.keys())

            if score == required_count:
                return rows, header_index, str(sheet.title), row_idx

    if best_score <= 0:
        raise ValueError(
            "Could not detect Excel header row. Place the course table header near the top of any sheet."
        )

    required_text = ", ".join(REQUIRED_COURSE_COLUMNS)
    found_text = ", ".join(best_columns) if best_columns else "none"
    raise ValueError(
        "Could not find all required course columns in Excel file. "
        f"Best match: sheet '{best_sheet}' row {best_row_number} with columns: {found_text}. "
        f"Required columns: {required_text}"
    )


def _school_payload_from_excel_bytes(content: bytes, filename: str) -> dict[str, Any]:
    if not content:
        raise ValueError("Uploaded file is empty.")
    if load_workbook is None:
        raise ValueError("openpyxl is not installed.")

    from io import BytesIO

    wb = load_workbook(filename=BytesIO(content), data_only=True)
    rows, header_index, selected_sheet_name, header_row_idx = _detect_course_header(wb)

    courses: list[dict[str, Any]] = []
    for row_offset, row in enumerate(rows[header_row_idx + 1 :], start=header_row_idx + 2):
        if row is None:
            continue
        values = list(row)
        course_code = str(_safe_cell(values, header_index.get("course_code")) or "").strip()
        if not course_code:
            continue
        course = {
            "course_code": course_code,
            "course_title": str(_safe_cell(values, header_index.get("course_title")) or "").strip(),
            "lecturer_id": str(_safe_cell(values, header_index.get("lecturer_id")) or "").strip(),
            "lectures_per_week": _safe_int(
                _safe_cell(values, header_index.get("lectures_per_week")),
                1,
                "lectures_per_week",
                row_offset,
            ),
            "min_working_days": _safe_int(
                _safe_cell(values, header_index.get("min_working_days")),
                1,
                "min_working_days",
                row_offset,
            ),
            "number_of_students": _safe_int(
                _safe_cell(values, header_index.get("number_of_students")),
                1,
                "number_of_students",
                row_offset,
            ),
            "department": str(_safe_cell(values, header_index.get("department")) or "").strip(),
            "level": str(_safe_cell(values, header_index.get("level")) or "").strip(),
        }
        courses.append(course)

    if not courses:
        raise ValueError(
            f"No course rows found in Excel file after header in sheet '{selected_sheet_name}'."
        )

    lecturers_map: dict[str, str] = {}
    rooms_map: dict[str, dict[str, Any]] = {}
    curricula_map: dict[str, set[str]] = {}
    unavailability: list[dict[str, Any]] = []

    for course in courses:
        lid = course["lecturer_id"]
        if lid:
            lecturers_map.setdefault(lid, lid)

    # Optional additional columns for richer payload inference
    if "lecturer_name" in header_index:
        for row in rows[header_row_idx + 1 :]:
            if row is None:
                continue
            values = list(row)
            lid = str(_safe_cell(values, header_index.get("lecturer_id")) or "").strip()
            if not lid:
                continue
            lname = str(_safe_cell(values, header_index.get("lecturer_name")) or "").strip()
            if lname:
                lecturers_map[lid] = lname

    if "room_id" in header_index:
        room_name_idx = header_index.get("room_name")
        room_capacity_idx = header_index.get("capacity")
        for row in rows[header_row_idx + 1 :]:
            if row is None:
                continue
            values = list(row)
            rid = str(_safe_cell(values, header_index.get("room_id")) or "").strip()
            if not rid:
                continue
            rname = (
                str(_safe_cell(values, room_name_idx) or "").strip() if room_name_idx is not None else rid
            )
            cap = _safe_int(_safe_cell(values, room_capacity_idx), 100, "capacity", header_row_idx + 2)
            rooms_map[rid] = {"room_id": rid, "room_name": rname or rid, "capacity": max(1, cap)}

    if "curriculum_id" in header_index:
        for row in rows[header_row_idx + 1 :]:
            if row is None:
                continue
            values = list(row)
            cid = str(_safe_cell(values, header_index.get("curriculum_id")) or "").strip()
            code = str(_safe_cell(values, header_index.get("course_code")) or "").strip()
            if cid and code:
                curricula_map.setdefault(cid, set()).add(code)

    if all(k in header_index for k in ("unavailability_day", "unavailability_period")):
        day_idx = header_index["unavailability_day"]
        per_idx = header_index["unavailability_period"]
        for row_offset, row in enumerate(rows[header_row_idx + 1 :], start=header_row_idx + 2):
            if row is None:
                continue
            values = list(row)
            code = str(_safe_cell(values, header_index.get("course_code")) or "").strip()
            if not code:
                continue
            day_val = _safe_cell(values, day_idx)
            per_val = _safe_cell(values, per_idx)
            if day_val is None or per_val is None:
                continue
            unavailability.append(
                {
                    "course_code": code,
                    "day": _safe_int(day_val, 0, "unavailability_day", row_offset),
                    "period": _safe_int(per_val, 0, "unavailability_period", row_offset),
                }
            )

    # Fallbacks when optional columns are absent
    if not rooms_map:
        rooms_map = {
            "R-DEFAULT-1": {"room_id": "R-DEFAULT-1", "room_name": "Main Hall", "capacity": 200},
            "R-DEFAULT-2": {"room_id": "R-DEFAULT-2", "room_name": "Lecture Room", "capacity": 120},
            "R-DEFAULT-3": {"room_id": "R-DEFAULT-3", "room_name": "Seminar Room", "capacity": 80},
        }
    if not curricula_map:
        for course in courses:
            dept = course["department"] or "General"
            level = course["level"] or "100"
            cid = f"CURR-{dept}-{level}".replace(" ", "_")
            curricula_map.setdefault(cid, set()).add(course["course_code"])

    curricula = [
        {
            "curriculum_id": cid,
            "curriculum_name": cid.replace("_", " "),
            "selected_course_codes": sorted(list(codes)),
        }
        for cid, codes in sorted(curricula_map.items())
    ]

    lecturers = [
        {"lecturer_id": lid, "lecturer_name": lname}
        for lid, lname in sorted(lecturers_map.items())
    ]
    rooms = [rooms_map[key] for key in sorted(rooms_map.keys())]

    return {
        "dataset_name": Path(filename).stem or "Excel_Dataset",
        "settings": {"number_of_days": 5, "periods_per_day": 6},
        "courses": courses,
        "lecturers": lecturers,
        "rooms": rooms,
        "curricula": curricula,
        "unavailability_constraints": unavailability,
    }


def _instance_from_school_data(
    payload: dict[str, Any],
) -> tuple[ITCInstance, dict[str, dict[str, str]]]:
    dataset_name = str(payload.get("dataset_name") or "School_Data").strip() or "School_Data"

    settings = payload.get("settings", {})
    if not isinstance(settings, dict):
        raise ValueError("settings must be an object.")
    nr_days = int(settings.get("number_of_days", 0))
    periods_per_day = int(settings.get("periods_per_day", 0))
    if nr_days <= 0 or periods_per_day <= 0:
        raise ValueError("number_of_days and periods_per_day must be positive.")

    courses_raw = payload.get("courses", [])
    rooms_raw = payload.get("rooms", [])
    curricula_raw = payload.get("curricula", [])
    unavailability_raw = payload.get("unavailability_constraints", [])
    lecturers_raw = payload.get("lecturers", [])

    if not isinstance(courses_raw, list) or not courses_raw:
        raise ValueError("courses must be a non-empty list.")
    if not isinstance(rooms_raw, list) or not rooms_raw:
        raise ValueError("rooms must be a non-empty list.")
    if not isinstance(curricula_raw, list):
        raise ValueError("curricula must be a list.")
    if not isinstance(unavailability_raw, list):
        raise ValueError("unavailability_constraints must be a list.")
    if not isinstance(lecturers_raw, list):
        raise ValueError("lecturers must be a list.")

    lecturer_ids = {
        str(entry.get("lecturer_id", "")).strip()
        for entry in lecturers_raw
        if isinstance(entry, dict)
    }

    course_meta: dict[str, dict[str, str]] = {}
    courses: list[Course] = []
    seen_course_codes: set[str] = set()
    for idx, entry in enumerate(courses_raw):
        if not isinstance(entry, dict):
            raise ValueError(f"courses[{idx}] must be an object.")
        course_code = str(entry.get("course_code", "")).strip()
        if not course_code:
            raise ValueError(f"courses[{idx}].course_code cannot be empty.")
        if course_code in seen_course_codes:
            raise ValueError(f"Duplicate course_code '{course_code}'.")
        seen_course_codes.add(course_code)

        lecturer_id = str(entry.get("lecturer_id", "")).strip()
        if not lecturer_id:
            raise ValueError(f"courses[{idx}].lecturer_id cannot be empty.")
        if lecturer_ids and lecturer_id not in lecturer_ids:
            raise ValueError(
                f"courses[{idx}].lecturer_id '{lecturer_id}' not found in lecturers list."
            )

        lectures_per_week = int(entry.get("lectures_per_week", 0))
        min_working_days = int(entry.get("min_working_days", 0))
        number_of_students = int(entry.get("number_of_students", 0))
        if lectures_per_week <= 0:
            raise ValueError(f"courses[{idx}].lectures_per_week must be positive.")
        if min_working_days <= 0:
            raise ValueError(f"courses[{idx}].min_working_days must be positive.")
        if number_of_students <= 0:
            raise ValueError(f"courses[{idx}].number_of_students must be positive.")

        courses.append(
            Course(
                course_id=course_code,
                teacher_id=lecturer_id,
                lectures_per_week=lectures_per_week,
                min_working_days=min_working_days,
                students=number_of_students,
            )
        )
        course_meta[course_code] = {
            "course_title": str(entry.get("course_title", "")).strip(),
            "lecturer_id": lecturer_id,
        }

    rooms: list[Room] = []
    seen_room_ids: set[str] = set()
    for idx, entry in enumerate(rooms_raw):
        if not isinstance(entry, dict):
            raise ValueError(f"rooms[{idx}] must be an object.")
        room_id = str(entry.get("room_id", "")).strip()
        if not room_id:
            raise ValueError(f"rooms[{idx}].room_id cannot be empty.")
        if room_id in seen_room_ids:
            raise ValueError(f"Duplicate room_id '{room_id}'.")
        seen_room_ids.add(room_id)

        capacity = int(entry.get("capacity", 0))
        if capacity <= 0:
            raise ValueError(f"rooms[{idx}].capacity must be positive.")
        rooms.append(Room(room_id=room_id, capacity=capacity))

    curricula: list[Curriculum] = []
    for idx, entry in enumerate(curricula_raw):
        if not isinstance(entry, dict):
            raise ValueError(f"curricula[{idx}] must be an object.")
        curriculum_id = str(entry.get("curriculum_id", "")).strip()
        if not curriculum_id:
            raise ValueError(f"curricula[{idx}].curriculum_id cannot be empty.")

        selected = entry.get("selected_course_codes", [])
        if not isinstance(selected, list):
            raise ValueError(f"curricula[{idx}].selected_course_codes must be a list.")
        selected_codes = [str(code).strip() for code in selected if str(code).strip()]
        for code in selected_codes:
            if code not in seen_course_codes:
                raise ValueError(
                    f"curricula[{idx}] references unknown course_code '{code}'."
                )
        curricula.append(
            Curriculum(curriculum_id=curriculum_id, course_ids=tuple(selected_codes))
        )

    unavailability_map: dict[str, set[tuple[int, int]]] = {}
    for idx, entry in enumerate(unavailability_raw):
        if not isinstance(entry, dict):
            raise ValueError(f"unavailability_constraints[{idx}] must be an object.")
        course_code = str(entry.get("course_code", "")).strip()
        if course_code not in seen_course_codes:
            raise ValueError(
                f"unavailability_constraints[{idx}].course_code '{course_code}' is unknown."
            )
        day = int(entry.get("day", -1))
        period = int(entry.get("period", -1))
        if day < 0 or day >= nr_days:
            raise ValueError(
                f"unavailability_constraints[{idx}].day must be in [0, {nr_days - 1}]."
            )
        if period < 0 or period >= periods_per_day:
            raise ValueError(
                "unavailability_constraints"
                f"[{idx}].period must be in [0, {periods_per_day - 1}]."
            )
        unavailability_map.setdefault(course_code, set()).add((day, period))

    unavailability = tuple(
        UnavailabilityConstraint(course_id=course_id, forbidden_slots=tuple(sorted(slots)))
        for course_id, slots in sorted(unavailability_map.items())
    )

    instance = ITCInstance(
        name=dataset_name,
        courses=tuple(courses),
        rooms=tuple(rooms),
        curricula=tuple(curricula),
        unavailability=unavailability,
        nr_days=nr_days,
        periods_per_day=periods_per_day,
    )
    return instance, course_meta


def _export_school_timetable_csv(
    state: TimetableState,
    output_path: Path,
    course_meta: dict[str, dict[str, str]],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "course_code",
                "course_title",
                "lecturer_id",
                "room_id",
                "day",
                "period",
                "day_label",
                "time_slot",
            ]
        )
        for assignment in state.assignments:
            info = course_meta.get(assignment.course_id, {})
            room = assignment.room_id if assignment.room_id is not None else "UNSCHEDULED"
            day = assignment.day if assignment.day is not None else "UNSCHEDULED"
            period = assignment.period if assignment.period is not None else "UNSCHEDULED"
            writer.writerow(
                [
                    assignment.course_id,
                    info.get("course_title", ""),
                    info.get("lecturer_id", ""),
                    room,
                    day,
                    period,
                    getDayName(assignment.day),
                    getTimeSlot(assignment.period),
                ]
            )


def _serialize_timetable_rows(
    state: TimetableState,
    course_meta: dict[str, dict[str, str]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for assignment in state.assignments:
        info = course_meta.get(assignment.course_id, {})
        rows.append(
            {
                "course_id": assignment.course_id,
                "course_title": info.get("course_title", ""),
                "lecturer_id": info.get("lecturer_id", ""),
                "room_id": assignment.room_id if assignment.room_id is not None else "UNSCHEDULED",
                "day": assignment.day,
                "period": assignment.period,
                "day_label": getDayName(assignment.day),
                "time_slot": getTimeSlot(assignment.period),
            }
        )
    return rows


def _course_meta_from_instance(instance: ITCInstance) -> dict[str, dict[str, str]]:
    meta: dict[str, dict[str, str]] = {}
    for course in instance.courses:
        meta[course.course_id] = {
            "course_title": course.course_id,
            "lecturer_id": course.teacher_id,
        }
    return meta
